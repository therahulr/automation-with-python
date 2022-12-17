from datetime import datetime

from api_auth import AuthAPI
import openpyxl

credential_path = 'conf/primary_credentials.json'
token_path = 'conf/token.json'


def main():
    query = 'from:(noreply@phonepe.com)'
    api_auth = AuthAPI(credential_path, token_path)
    service = api_auth.get_service()
    messages = api_auth.get_messages(service, query)

    for msg in range(len(messages)):
        if messages[msg].startswith('PhonePe'):
            print(f"Message >> {msg + 1} >> {messages[msg]}")
            with open("phonepe_emails.txt", "a", encoding="utf-8") as f:
                f.write(messages[msg] + "\n")

    # Write to Excel file
    excel_file = "phonepe_emails.xlsx"
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    # Fill headers
    headers = ["Date", "Transaction Type", "Amount", "Name", "Bank Name"]
    for col in range(len(headers)):
        sheet.cell(row=1, column=col + 1).value = headers[col]

    # Fill data
    for msgs in range(len(messages)):
        if messages[msgs].startswith('PhonePe'):
            if "paid to" in messages[msgs].lower():
                date = datetime.strptime(messages[msgs].split("Paid to")[0].strip().split("PhonePe")[1].strip(),
                                         "%b %d, %Y").strftime(
                    "%d-%m-%Y")
                name = messages[msgs].split("Paid to")[1].split("₹")[0].strip()
                amount = messages[msgs].split("Paid to")[1].split("₹")[1].split("Txn. ID")[0].strip()
                debited_from = messages[msgs].split("Debited from")[1].split("Bank")[0].split(":")[1].strip() + " Bank"
                sheet.cell(row=msgs + 2, column=1).value = date
                sheet.cell(row=msgs + 2, column=2).value = "Debit"
                sheet.cell(row=msgs + 2, column=3).value = amount
                sheet.cell(row=msgs + 2, column=4).value = name
                sheet.cell(row=msgs + 2, column=5).value = debited_from

            if "received from" in messages[msgs].lower():
                date = datetime.strptime(messages[msgs].split("Received from")[0].strip().split("PhonePe")[1].strip(),
                                         "%b %d, %Y").strftime("%d-%m-%Y")
                name = messages[msgs].split("Received from")[1].split("₹")[0].strip()
                amount = messages[msgs].split("Received from")[1].split("₹")[1].split("Txn. ID")[0].strip()
                credited_to = messages[msgs].split("Credited to")[1].split("Bank")[0].split(":")[1].strip() + " Bank"
                sheet.cell(row=msgs + 2, column=1).value = date
                sheet.cell(row=msgs + 2, column=2).value = "Credit"
                sheet.cell(row=msgs + 2, column=3).value = amount
                sheet.cell(row=msgs + 2, column=4).value = name
                sheet.cell(row=msgs + 2, column=5).value = credited_to
    wb.save(excel_file)


if __name__ == '__main__':
    main()
